from datetime import date, datetime, timedelta
from collections import defaultdict
from io import BytesIO
from flask import Blueprint, render_template, flash, request, jsonify, send_file
from flask_login import login_required, current_user
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..forms import DashboardFilterForm
from ..models import ExamLog, Doctor, Category, db

bp = Blueprint("dashboard", __name__)


def get_korean_date_string(d):
    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
    return f"{d.year}년 {d.month}월 {d.day}일 ({weekdays[d.weekday()]})"


def _resolve_dates(form: DashboardFilterForm):
    # Priority 1: Query Parameters (Navigation)
    q_start = request.args.get('start_date')
    q_end = request.args.get('end_date')
    if q_start and q_end:
        try:
            return datetime.strptime(q_start, '%Y-%m-%d').date(), datetime.strptime(q_end, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Priority 2: Form Data
    today = date.today()
    if form.period.data == "today":
        return today, today
    if form.period.data == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end
    if form.period.data == "month":
        start = today.replace(day=1)
        if start.month == 12:
            next_month = start.replace(year=start.year + 1, month=1, day=1)
        else:
            next_month = start.replace(month=start.month + 1, day=1)
        end = next_month - timedelta(days=1)
        return start, end
    return form.start_date.data, form.end_date.data


def _aggregate(start, end):
    # Generate list of dates
    delta = end - start
    dates = [start + timedelta(days=i) for i in range(delta.days + 1)]
    
    start_str = start.strftime('%Y-%m-%d')
    end_str = end.strftime('%Y-%m-%d')

    # Query ExamLog joined with Category and Doctor
    query = db.session.query(ExamLog, Category, Doctor).join(
        Category, ExamLog.category_id == Category.category_id
    ).join(
        Doctor, ExamLog.doctor_id == Doctor.doctor_id
    ).filter(
        ExamLog.exam_date >= start_str, 
        ExamLog.exam_date <= end_str
    )
    
    records = query.all()

    # Structure: data[doctor_name][procedure_code][metric_suffix] = count
    # e.g. matrix['M1']['ENDO']['PUB_S'] = 5
    
    matrix = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    summary = defaultdict(int)
    doc_stats = defaultdict(lambda: {'chk_total': 0, 'out_total': 0})

    for log, cat, doc in records:
        impact = 1 if log.action_type == 'INSERT' else -1
        
        # Grid Mapping (Procedure Categories)
        # metric_code format: {PROC}_{SUFFIX} e.g. ENDO_PUB_S
        if '_' in cat.metric_code:
            try:
                # metric_code "ENDO_PUB_S" -> proc="ENDO", suffix="PUB_S"
                parts = cat.metric_code.split('_', 1)
                proc_code = parts[0]
                metric_suffix = parts[1]
                
                matrix[doc.doctor_name][proc_code][metric_suffix] += impact
                
                # Doctor Stats (Row Totals for Div Column)
                # Checkup Row (검): Includes PUB (Public) and CHK (Checkup)
                if metric_suffix in ['CHK_S', 'CHK_G', 'PUB_S', 'PUB_G']:
                    doc_stats[doc.doctor_name]['chk_total'] += impact
                
                # Outpatient Row (외): Includes OUT (Outpatient) and INP (Inpatient/Ward)
                if metric_suffix in ['OUT_S', 'OUT_G', 'INP_S', 'INP_G']:
                    doc_stats[doc.doctor_name]['out_total'] += impact
                
                # Global Summaries
                # General = Ends with _G
                # Sedation = Ends with _S
                
                if metric_suffix.endswith('_G'):
                    summary[f'{proc_code}_general'] += impact
                    summary['general_total'] += impact
                
                if metric_suffix.endswith('_S'):
                    summary[f'{proc_code}_sedation'] += impact
                    summary['sedation_total'] += impact

                summary[proc_code] += impact # Total per procedure type

                # Specific Doctor/Metric Summaries
                # "외수" maps to OUT_S (Outpatient Sedation)
                if metric_suffix == 'OUT_S':
                    if doc.doctor_name in ['M1', 'M2', 'M4']:
                        summary[f'{doc.doctor_name}_OUT_S'] += impact
                
                # "병수" maps to INP_S (Inpatient Sedation)
                if metric_suffix == 'INP_S':
                    summary['INP_S_total'] += impact
                    
                # HC Total: Sum of PUB (Public) and CHK (Checkup)
                # "공수 검수 를 HC 총합 으로 구성"
                if metric_suffix.startswith('PUB') or metric_suffix.startswith('CHK'):
                    summary['HC_total'] += impact

            except ValueError:
                pass

    summary["overall"] = summary.get("general_total", 0) + summary.get("sedation_total", 0)
    
    return matrix, summary, dates, doc_stats


@bp.route("/")
@login_required
def index():
    form = DashboardFilterForm()
    if form.validate_on_submit():
        start, end = _resolve_dates(form)
    else:
        start, end = _resolve_dates(form)
    
    matrix, summary, dates, doc_stats = _aggregate(start, end)
    # Order by ID to preserve the specific insertion order requested (e.g. M1...M30)
    doctors = Doctor.query.filter_by(is_active=1).order_by(Doctor.doctor_id).all()
    
    # Procedures list for Column Headers
    procedures = ['ENDO', 'COLON', 'ERCP', 'SIG', 'PEG', 'CESD', 'GESD']
    
    # Configuration for partial read-only doctors (Allow-List approach)
    # Doctors NOT in this list have access to ALL procedures.
    # Doctors IN this list have access ONLY to the specified procedures.
    allow_config = {
        'M12': ['ENDO', 'COLON', 'ERCP', 'SIG', 'PEG'],
        'FM2': ['ENDO', 'COLON', 'ERCP', 'SIG', 'PEG'],
        'FM3': ['ENDO', 'COLON', 'ERCP', 'SIG', 'PEG'],
        'M30': ['ENDO', 'COLON', 'ERCP', 'SIG', 'PEG'],
        'FM1': ['ENDO', 'COLON']
    }

    # Calculate navigation dates
    prev_start = (start - timedelta(days=1)).strftime('%Y-%m-%d')
    prev_end = (end - timedelta(days=1)).strftime('%Y-%m-%d')
    next_start = (start + timedelta(days=1)).strftime('%Y-%m-%d')
    next_end = (end + timedelta(days=1)).strftime('%Y-%m-%d')

    # Korean Date String
    formatted_date = get_korean_date_string(start)

    return render_template(
        "dashboard.html",
        form=form,
        start=start,
        end=end,
        formatted_date=formatted_date,
        prev_start=prev_start,
        prev_end=prev_end,
        next_start=next_start,
        next_end=next_end,
        matrix=matrix,
        summary=summary,
        doc_stats=doc_stats,
        dates=dates,
        doctors=doctors,
        procedures=procedures,
        allow_config=allow_config,
        patient_groups=[],
        procedure_types=[],
    )



@bp.route("/api/get_details", methods=["GET"])
@login_required
def get_details():
    doctor_name = request.args.get("doctor_name")
    procedure = request.args.get("procedure") 
    date_str = request.args.get("date")

    if not date_str:
        date_str = date.today().strftime('%Y-%m-%d')
    
    doctor = Doctor.query.filter_by(doctor_name=doctor_name).first()
    if not doctor:
        return jsonify({"error": "Doctor not found"}), 404

    # Fetch logs for this doctor/date/procedure
    # We filter by Category group_code == procedure (e.g. 'ENDO')
    
    logs = db.session.query(ExamLog, Category).join(
        Category, ExamLog.category_id == Category.category_id
    ).filter(
        ExamLog.doctor_id == doctor.doctor_id,
        ExamLog.exam_date == date_str,
        Category.group_code == procedure
    ).all()
    
    counts = defaultdict(int)
    for log, cat in logs:
        impact = 1 if log.action_type == 'INSERT' else -1
        # metric_code example "ENDO_PUB_S" -> "PUB_S"
        if '_' in cat.metric_code:
            suffix = cat.metric_code.split('_', 1)[1]
            counts[suffix] += impact
        
    return jsonify(counts)


@bp.route("/api/update_log", methods=["POST"])
@login_required
def update_log():
    data = request.get_json()
    print(f"DEBUG: update_log received: {data}") # Debug logging
    
    doctor_name = data.get("doctor_name")
    date_str = data.get("date")
    # metric_code now expects full code e.g. "ENDO_PUB_S"
    metric_code = data.get("metric_code") 
    action = data.get("action") 
    value = data.get("value")
    
    if not date_str:
        date_str = date.today().strftime('%Y-%m-%d')
        print(f"DEBUG: No date provided, defaulting to {date_str}")
        
    doctor = Doctor.query.filter_by(doctor_name=doctor_name).first()
    if not doctor:
        print(f"ERROR: Doctor {doctor_name} not found")
        return jsonify({"error": "Doctor not found"}), 404
    
    category = Category.query.filter_by(metric_code=metric_code).first()
    if not category:
        return jsonify({"error": f"Category {metric_code} not found"}), 404
        
    # Logic for SETTING value absolute
    # If action is 'set', we need to adjust the log to match the new value.
    # Simplified approach: Delete existing logs for this metric/day/doc and insert new count?
    # Or just calculate difference?
    # Current log system is Event Sourcing (INSERT/DELETE).
    # To set specific value X:
    # 1. Get current sum.
    # 2. Add/Subtract difference.
    
    if action == 'set' and value is not None:
        # Calculate current qty
        current_logs = ExamLog.query.filter_by(
            doctor_id=doctor.doctor_id,
            category_id=category.category_id,
            exam_date=date_str
        ).all()
        current_qty = sum(1 if l.action_type == 'INSERT' else -1 for l in current_logs)
        
        diff = int(value) - current_qty
        
        if diff > 0:
            for _ in range(diff):
                db.session.add(ExamLog(
                    exam_date=date_str, doctor_id=doctor.doctor_id, category_id=category.category_id,
                    created_by=current_user.user_id, action_type='INSERT'
                ))
        elif diff < 0:
            for _ in range(abs(diff)):
                db.session.add(ExamLog(
                    exam_date=date_str, doctor_id=doctor.doctor_id, category_id=category.category_id,
                    created_by=current_user.user_id, action_type='DELETE'
                ))
    else:
        # Lagacy support for increment/decrement if needed, though frontend now sends 'set'
        action_type = 'INSERT' if action == 'increment' else 'DELETE'
        db.session.add(ExamLog(
            exam_date=date_str,
            doctor_id=doctor.doctor_id,
            category_id=category.category_id,
            created_by=current_user.user_id,
            action_type=action_type
        ))
    
    db.session.commit()
    
    return jsonify({"success": True})

@bp.route("/dashboard", methods=["GET", "POST"])
@login_required
def dashboard_view():
    return index()


@bp.route("/export_excel")
@login_required
def export_excel():
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')
    
    if not start_str or not end_str:
         start = date.today()
         end = date.today()
    else:
        start = datetime.strptime(start_str, '%Y-%m-%d').date()
        end = datetime.strptime(end_str, '%Y-%m-%d').date()

    matrix, summary, dates, doc_stats = _aggregate(start, end)
    doctors = Doctor.query.filter_by(is_active=1).order_by(Doctor.doctor_id).all()
    procedures = ['ENDO', 'COLON', 'ERCP', 'SIG', 'PEG', 'CESD', 'GESD']

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Endo Dashboard"

    # Header Row
    ws['A1'] = "Doctor"
    ws['B1'] = "Div"
    
    col_idx = 3
    for proc in procedures:
        ws.cell(row=1, column=col_idx, value=proc)
        col_idx += 1
    
    # Data Rows
    row_idx = 2
    for doc in doctors:
        dname = doc.doctor_name
        
        # Checkup Row
        ws.cell(row=row_idx, column=1, value=dname)
        chk_total = doc_stats[dname]['chk_total']
        ws.cell(row=row_idx, column=2, value=f"검({chk_total})")
        
        c = 3
        for proc in procedures:
            val_str = f"공:{matrix[dname][proc]['PUB_S'] or 0} / 검:{matrix[dname][proc]['CHK_S'] or 0}"
            ws.cell(row=row_idx, column=c, value=val_str)
            c += 1
        row_idx += 1
        
        # Outpatient Row (Similar to HTML logic)
        out_total = doc_stats[dname]['out_total']
        ws.cell(row=row_idx, column=2, value=f"외({out_total})")
        c = 3
        for proc in procedures:
            val_str = f"외:{matrix[dname][proc]['OUT_S'] or 0} / 병:{matrix[dname][proc]['INP_S'] or 0}"
            ws.cell(row=row_idx, column=c, value=val_str)
            c += 1
        row_idx += 2 # Skip a line or just strict rows? Let's keep it compact.

    # Footer Summaries (Simplified for Excel)
    ws.cell(row=row_idx, column=1, value="SUMMARY")
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="일반 합계")
    c = 3
    for proc in procedures:
        ws.cell(row=row_idx, column=c, value=summary[proc + '_general'])
        c += 1
    row_idx += 1

    ws.cell(row=row_idx, column=1, value="수면 합계")
    c = 3
    for proc in procedures:
        ws.cell(row=row_idx, column=c, value=summary[proc + '_sedation'])
        c += 1
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="전체 총계")
    c = 3
    for proc in procedures:
        ws.cell(row=row_idx, column=c, value=summary[proc])
        c += 1
    row_idx += 2

    ws.cell(row=row_idx, column=1, value=f"HC Total: {summary['HC_total']}")
    ws.cell(row=row_idx+1, column=1, value=f"Overall Total: {summary['overall']}")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Endo_Dashboard_{start.strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)
